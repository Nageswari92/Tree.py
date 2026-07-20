import streamlit as st
import pandas as pd
import io
from openpyxl.styles import PatternFill 

# 1. Page Configuration (Web App UI Layout)
st.set_page_config(
    page_title="BAVA TECH - BOM Tree Processor", 
    page_icon="⚙️", 
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for modern styling
st.markdown("""
    <style>
    .main { background-color: #f8f9fa; }
    .stButton>button { width: 100%; border-radius: 8px; font-weight: bold; }
    h1 { color: #1E3A8A; }
    </style>
    """, unsafe_allow_html=True)

# 2. Sidebar Navigation / Instructions
with st.sidebar:
    st.image("https://img.icons8.com/fluent/96/000000/tree-structure.png", width=80)
    st.title("BAVA TECH")
    st.caption("BOM Management Suite")
    st.markdown("---")
    st.subheader("💡 How to use:")
    st.write("1. Prepare an Excel file with a sheet named **'BOM'**.")
    st.write("2. Ensure it has **'Level'** and **'Part Number'** columns.")
    st.write("3. Upload the file and download your processed Tree hierarchy instantly.")
    st.markdown("---")
    st.caption("v2.1.0 | Core Engine")

# 3. Main Web App Dashboard
st.title("⚙️ BAVA TECH - BOM Tree Creation Process")
st.write("Efficiently transform raw, multi-level Bill of Materials (BOM) sheets into structured engineering parent-child tree hierarchies.")
st.markdown("---")

# Layout columns for uploading and previewing
col1, col2 = st.columns([1, 2])

with col1:
    st.subheader("📥 Upload Workspace")
    uploaded_file = st.file_uploader(
        "Drag and drop your BOM file here", 
        type=["xls", "xlsx"],
        help="Supports .xls and .xlsx formats with a 'BOM' sheet tab."
    )

with col2:
    if uploaded_file is not None:
        customer_name = uploaded_file.name.rsplit('.', 1)[0]
        st.subheader("📋 Active Session")
        
        st.info(f"**Filename:** {uploaded_file.name} | **Identified Customer:** {customer_name}")
        
        try:
            # Core processing logic
            df_bom = pd.read_excel(uploaded_file, sheet_name="BOM")
            
            if "Level" not in df_bom.columns or "Part Number" not in df_bom.columns:
                st.error("❌ Structure Error: Missing 'Level' or 'Part Number' columns in the 'BOM' sheet.")
            else:
                with st.spinner("Executing BOM Tree Creation Process... Please wait..."):
                    tree_data = []
                    total_rows = len(df_bom)

                    for index, row in df_bom.iterrows():
                        level = row["Level"]
                        part_number = str(row["Part Number"]).strip() 
                        
                        tree_row = {
                            "Customer": customer_name if index == 0 else None,
                            "Top Level": None,
                            "Level 2": None,
                            "Level 3": None,
                            "Level 4": None,
                            "Level 5": None
                        }
                        
                        has_child = False
                        if level in [0, 1, 2, 3, 4]:
                            for j in range(index + 1, total_rows):
                                if df_bom.iloc[j]["Level"] == level + 1:
                                    has_child = True
                                    break
                                elif df_bom.iloc[j]["Level"] <= level:
                                    break

                        if level == 0:
                            tree_row["Top Level"] = part_number
                            tree_data.append(tree_row)
                        elif level == 1 and has_child:
                            tree_row["Level 2"] = part_number
                            tree_data.append(tree_row)
                        elif level == 2 and has_child:
                            tree_row["Level 3"] = part_number
                            tree_data.append(tree_row)
                        elif level == 3 and has_child:
                            tree_row["Level 4"] = part_number
                            tree_data.append(tree_row)
                        elif level == 4 and has_child:
                            tree_row["Level 5"] = part_number
                            tree_data.append(tree_row)

                    df_tree = pd.DataFrame(tree_data)
                    df_tree = df_tree[["Customer", "Top Level", "Level 2", "Level 3", "Level 4", "Level 5"]]
                    
                    # File conversion for user download with openpyxl styling
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df_tree.to_excel(writer, index=False)
                        
                        workbook = writer.book
                        worksheet = writer.sheets['Sheet1']
                        
                        # Color definition for "Orange, Accent 2, Lighter 80%" (Hex: FCE4D6)
                        orange_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
                        
                        # "0030" removed, "0300" remains active
                        prefixes = ("0042", "0040", "0043", "0243", "0015", "0300")
                        
                        # Iterate cells (skip header row 1)
                        for row in worksheet.iter_rows(min_row=2, min_col=1, max_col=6):
                            for cell in row:
                                if cell.value is not None:
                                    val_str = str(cell.value).strip()
                                    if val_str.startswith(prefixes):
                                        cell.fill = orange_fill
                                        
                    processed_data = output.getvalue()
                    
                    st.success("🎉 BoM_Tree Creation & Highlighting Process Completed Successfully!")
                    
                    output_filename = f"BoM_Tree_{customer_name}.xlsx"
                    st.download_button(
                        label="⚡ Download Highlighted BoM_Tree Excel",
                        data=processed_data,
                        file_name=output_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary"
                    )

        except Exception as e:
            st.error(f"❌ Processing Interrupted: {e}")
    else:
        st.warning("💤 Awaiting document upload from the panel on the left.")

# 4. Live Data View Area
if uploaded_file is not None and 'df_tree' in locals():
    st.markdown("---")
    st.subheader("📊 Created BoM_Tree Data Explorer")
    st.dataframe(df_tree, width="stretch")
